#!/usr/bin/env python3
"""주요품목 수출입 엑셀 → data/exports.json 변환.

관세청 무역통계 기반으로 품목별(시트별) 월별 수출입 실적을 관리하는 엑셀에서
원자료 블록(년월 / 금액(달러) / 금액(원화) / 중량(Kg))만 추출한다.

사용법:
  pip install openpyxl
  python3 scripts/ingest-exports-xlsx.py <수출입엑셀.xlsx>

시트 규칙:
  - 1행: 관련 기업, 2행: HS코드 (첫 번째 비어있지 않은 셀)
  - 상단 헤더 몇 행 안에서 '년월' 셀 오른쪽에 '금액…' 헤더가 오는 열을 원자료 블록으로 인식
  - '2022년01월' 형식의 행만 월별 데이터로 수집 (10일 단위 잠정치/연간 합계 행은 제외)
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

MONTH_RE = re.compile(r"^\s*(\d{4})년\s*(\d{2})월\s*$")
SKIP_SHEETS = {"반도체 수출 판가 증가율"}  # 파생/요약 시트
HEADER_WORDS = {"년월", "금액(달러)", "금액(원화)", "중량(Kg)"}
# 총수출 시트는 관세청 발표 단위(천달러)로 입력되어 있어 달러로 환산한다.
KUSD_SHEETS = {"총수출"}


def looks_like_hs(v) -> bool:
    s = str(v).strip()
    digits = sum(c.isdigit() for c in s)
    return digits >= 4 and digits / max(len(s), 1) > 0.5


def sheet_meta(rows, title):
    """상단 1~3행에서 HS코드(숫자 위주 값)와 관련 기업(텍스트)을 찾는다."""
    hs = None
    companies = None
    for row in rows[:3]:
        for v in row or ():
            if v is None:
                continue
            s = str(v).strip()
            if not s or s in HEADER_WORDS:
                continue
            if looks_like_hs(v):
                if hs is None:
                    hs = s
            elif isinstance(v, str) and companies is None and s != title:
                companies = s
    return hs, companies


def find_anchor(rows):
    """원자료 블록의 ('년월' 헤더 행, 열) 위치를 찾는다."""
    for ri in range(min(8, len(rows))):
        row = rows[ri] or ()
        for ci, v in enumerate(row):
            if isinstance(v, str) and v.strip() == "년월" and ci + 1 < len(row):
                nxt = row[ci + 1]
                if isinstance(nxt, str) and "금액" in nxt:
                    return ri, ci
    return None


def num(v):
    if v is None or isinstance(v, str):
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return f


def main():
    if len(sys.argv) < 2:
        sys.exit("usage: ingest-exports-xlsx.py <xlsx>")
    src = Path(sys.argv[1])
    out = Path(__file__).resolve().parent.parent / "data" / "exports.json"

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    items = []
    skipped = []
    for name in wb.sheetnames:
        title = name.strip()
        if title in SKIP_SHEETS or name in SKIP_SHEETS:
            continue
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        anchor = find_anchor(rows)
        if not anchor:
            skipped.append(title)
            continue
        ri, ci = anchor
        hs, companies = sheet_meta(rows, title)
        scale = 1000.0 if title in KUSD_SHEETS else 1.0
        by_month = {}
        for r in rows[ri + 1:]:
            if not r or ci >= len(r) or r[ci] is None:
                continue
            m = MONTH_RE.match(str(r[ci]))
            if not m:
                continue
            usd = num(r[ci + 1]) if ci + 1 < len(r) else None
            krw = num(r[ci + 2]) if ci + 2 < len(r) else None
            kg = num(r[ci + 3]) if ci + 3 < len(r) else None
            if usd is None and krw is None and kg is None:
                continue
            if usd is not None:
                usd *= scale
            row = [f"{m.group(1)}-{m.group(2)}", usd, krw, kg]
            prev = by_month.get(row[0])
            # 같은 달이 두 번 나오면 값이 더 채워진 행을 유지
            if prev is None or sum(v is not None for v in row[1:]) > sum(v is not None for v in prev[1:]):
                by_month[row[0]] = row
        if not by_month:
            skipped.append(title)
            continue
        series = [by_month[k] for k in sorted(by_month)]
        if title in KUSD_SHEETS:
            hs, companies = None, None  # 국가 전체 실적: 시트에 남은 품목 메타는 무시
        items.append({
            "name": title,
            "hs": hs,
            "companies": companies,
            "rows": series,
        })

    # 기존 데이터와 병합: 같은 품목·같은 달은 엑셀 값이 이기고,
    # 엑셀에 없는 과거 달(API 백필분 등)은 유지한다.
    try:
        prev = json.loads(out.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        prev = {"items": []}
    prev_by_name = {it["name"]: it for it in prev.get("items", [])}
    for it in items:
        old = prev_by_name.get(it["name"])
        if not old:
            continue
        merged = {r[0]: r for r in old.get("rows", [])}
        for r in it["rows"]:
            merged[r[0]] = r
        it["rows"] = [merged[k] for k in sorted(merged)]
    names = {it["name"] for it in items}
    items += [it for it in prev.get("items", []) if it["name"] not in names]

    data = {
        "source": src.name,
        "items": items,
    }
    out.write_text(json.dumps(data, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    months = sorted({r[0] for it in items for r in it["rows"]})
    print(f"items: {len(items)}  months: {months[0]}..{months[-1]}  -> {out} ({out.stat().st_size/1e6:.2f} MB)")
    if skipped:
        print("skipped:", ", ".join(skipped))


if __name__ == "__main__":
    main()
